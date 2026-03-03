#!/usr/bin/env python3
"""
Compare `Sheet1` AI outputs against the `QA` sheet expectations and write a report.

Ground truth rules:
- Expected Year: `QA` sheet column `AI Year`
- Expected Make/Model/Trim: use `Manual ...` if non-empty; otherwise use `AI ...`

Outputs:
- Prints summary accuracy to stdout
- Writes a new workbook with an added `Accuracy_Report` sheet

Usage:
  python qa_accuracy_report.py "Boat mart demo.filled.xlsx" "Boat mart demo.filled.accuracy.xlsx"
"""

from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from typing import Any, Dict, Tuple

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def _header_map(ws) -> Dict[str, int]:
    m: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, str) and v.strip():
            m[v.strip()] = c
    return m


def _as_text(x: Any) -> str:
    if x is None:
        return ""
    if isinstance(x, bool):
        return str(x)
    if isinstance(x, int):
        return str(x)
    if isinstance(x, float):
        return str(int(x)) if x.is_integer() else str(x)
    return str(x).strip()


def _canon(x: Any) -> str:
    return re.sub(r"\s+", " ", _as_text(x)).strip()


def _canon_ci(x: Any) -> str:
    return _canon(x).lower()


_MISSING_TOKENS = {"", "n/a", "na", "none", "null"}


def _is_missing(x: Any) -> bool:
    return _canon_ci(x) in _MISSING_TOKENS


def _relaxed_equiv(a: Any, b: Any, field: str) -> bool:
    """
    Relaxed equivalence used for analysis:
    - case-insensitive match
    - treat Not found ~= blank
    - treat Standard ~= blank / Not found for Trim
    - punctuation-insensitive for Make/Model/Trim
    """
    a0 = _canon_ci(a)
    b0 = _canon_ci(b)
    if a0 == b0:
        return True

    nf = {"not found"}
    if (a0 in nf and b0 in _MISSING_TOKENS) or (b0 in nf and a0 in _MISSING_TOKENS):
        return True

    if field == "Trim":
        std = {"standard"}
        if (a0 in std and (b0 in _MISSING_TOKENS or b0 in nf)) or (b0 in std and (a0 in _MISSING_TOKENS or a0 in nf)):
            return True

    if field in {"Make", "Model", "Trim"}:
        ra = re.sub(r"[^a-z0-9]+", "", a0)
        rb = re.sub(r"[^a-z0-9]+", "", b0)
        if ra and ra == rb:
            return True

    return False


@dataclass(frozen=True)
class CompareRow:
    row_sheet1: int
    scraped_name: str
    qa_status: str
    actual_year: str
    expected_year: str
    actual_make: str
    expected_make: str
    actual_model: str
    expected_model: str
    actual_trim: str
    expected_trim: str
    strict_year: bool
    strict_make: bool
    strict_model: bool
    strict_trim: bool
    strict_all: bool
    relaxed_year: bool
    relaxed_make: bool
    relaxed_model: bool
    relaxed_trim: bool
    relaxed_all: bool


def _pick_expected(wsq, hq: Dict[str, int], qa_row: int, field: str) -> Any:
    if field == "Year":
        return wsq.cell(row=qa_row, column=hq["AI Year"]).value
    if field == "Make":
        manual = wsq.cell(row=qa_row, column=hq["Manual Make"]).value
        return manual if _as_text(manual) != "" else wsq.cell(row=qa_row, column=hq["AI Make"]).value
    if field == "Model":
        manual = wsq.cell(row=qa_row, column=hq["Manual Model"]).value
        return manual if _as_text(manual) != "" else wsq.cell(row=qa_row, column=hq["AI Model"]).value
    if field == "Trim":
        manual = wsq.cell(row=qa_row, column=hq["Manual Trim"]).value
        return manual if _as_text(manual) != "" else wsq.cell(row=qa_row, column=hq["AI Trim"]).value
    raise KeyError(field)


def _get_actual(ws1, h1: Dict[str, int], sheet1_row: int, field: str) -> Any:
    if field == "Year":
        return ws1.cell(row=sheet1_row, column=h1["AI_Year"]).value
    if field == "Make":
        return ws1.cell(row=sheet1_row, column=h1["AI_Make"]).value
    if field == "Model":
        return ws1.cell(row=sheet1_row, column=h1["AI_Model"]).value
    if field == "Trim":
        return ws1.cell(row=sheet1_row, column=h1["AI_Trim"]).value
    raise KeyError(field)


def compare_workbook(input_path: str) -> Tuple[list[CompareRow], Dict[str, float]]:
    wb = load_workbook(input_path, data_only=True)
    if "Sheet1" not in wb.sheetnames or "QA" not in wb.sheetnames:
        raise RuntimeError(f"Workbook must contain sheets 'Sheet1' and 'QA'. Found: {wb.sheetnames}")

    ws1 = wb["Sheet1"]
    wsq = wb["QA"]
    h1 = _header_map(ws1)
    hq = _header_map(wsq)

    for req in ["Scraped Name", "AI Year", "AI Make", "AI Model", "AI Trim", "QA", "Manual Make", "Manual Model", "Manual Trim"]:
        if req not in hq:
            raise RuntimeError(f"QA sheet missing required column: {req!r}. Found: {list(hq.keys())}")
    for req in ["Scraped Name", "AI_Year", "AI_Make", "AI_Model", "AI_Trim"]:
        if req not in h1:
            raise RuntimeError(f"Sheet1 missing required column: {req!r}. Found: {list(h1.keys())}")

    qa_by_name: Dict[str, int] = {}
    for r in range(2, wsq.max_row + 1):
        nm = wsq.cell(row=r, column=hq["Scraped Name"]).value
        if isinstance(nm, str) and nm.strip():
            qa_by_name.setdefault(nm.strip(), r)

    rows: list[CompareRow] = []
    fields = ["Year", "Make", "Model", "Trim"]

    counts_strict = {f: 0 for f in fields}
    counts_relaxed = {f: 0 for f in fields}
    total = 0
    strict_all = 0
    relaxed_all = 0

    for r in range(2, ws1.max_row + 1):
        nm = ws1.cell(row=r, column=h1["Scraped Name"]).value
        if not isinstance(nm, str) or not nm.strip():
            continue
        name = nm.strip()
        qa_row = qa_by_name.get(name)
        if not qa_row:
            continue

        total += 1
        qa_status = _as_text(wsq.cell(row=qa_row, column=hq["QA"]).value) or ""

        actual = {f: _get_actual(ws1, h1, r, f) for f in fields}
        expected = {f: _pick_expected(wsq, hq, qa_row, f) for f in fields}

        # strict checks
        strict = {
            "Year": _canon(actual["Year"]) == _canon(expected["Year"]),
            "Make": _canon_ci(actual["Make"]) == _canon_ci(expected["Make"]),
            "Model": _canon_ci(actual["Model"]) == _canon_ci(expected["Model"]),
            "Trim": _canon_ci(actual["Trim"]) == _canon_ci(expected["Trim"]),
        }

        relaxed = {f: _relaxed_equiv(actual[f], expected[f], f) for f in fields}

        for f in fields:
            if strict[f]:
                counts_strict[f] += 1
            if relaxed[f]:
                counts_relaxed[f] += 1

        strict_row_all = all(strict[f] for f in fields)
        relaxed_row_all = all(relaxed[f] for f in fields)
        if strict_row_all:
            strict_all += 1
        if relaxed_row_all:
            relaxed_all += 1

        rows.append(
            CompareRow(
                row_sheet1=r,
                scraped_name=name,
                qa_status=qa_status,
                actual_year=_as_text(actual["Year"]),
                expected_year=_as_text(expected["Year"]),
                actual_make=_as_text(actual["Make"]),
                expected_make=_as_text(expected["Make"]),
                actual_model=_as_text(actual["Model"]),
                expected_model=_as_text(expected["Model"]),
                actual_trim=_as_text(actual["Trim"]),
                expected_trim=_as_text(expected["Trim"]),
                strict_year=strict["Year"],
                strict_make=strict["Make"],
                strict_model=strict["Model"],
                strict_trim=strict["Trim"],
                strict_all=strict_row_all,
                relaxed_year=relaxed["Year"],
                relaxed_make=relaxed["Make"],
                relaxed_model=relaxed["Model"],
                relaxed_trim=relaxed["Trim"],
                relaxed_all=relaxed_row_all,
            )
        )

    def _rate(n: int) -> float:
        return (n / total) if total else 0.0

    summary = {
        "total_rows": float(total),
        "strict_year": _rate(counts_strict["Year"]),
        "strict_make": _rate(counts_strict["Make"]),
        "strict_model": _rate(counts_strict["Model"]),
        "strict_trim": _rate(counts_strict["Trim"]),
        "strict_all": _rate(strict_all),
        "relaxed_year": _rate(counts_relaxed["Year"]),
        "relaxed_make": _rate(counts_relaxed["Make"]),
        "relaxed_model": _rate(counts_relaxed["Model"]),
        "relaxed_trim": _rate(counts_relaxed["Trim"]),
        "relaxed_all": _rate(relaxed_all),
    }

    return rows, summary


def write_report(input_path: str, output_path: str, rows: list[CompareRow], summary: Dict[str, float]) -> None:
    wb = load_workbook(input_path)

    # remove existing report sheet if present
    report_name = "Accuracy_Report"
    if report_name in wb.sheetnames:
        wb.remove(wb[report_name])
    ws = wb.create_sheet(report_name)

    # styles
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bad_fill = PatternFill("solid", fgColor="FFC7CE")  # light red
    good_fill = PatternFill("solid", fgColor="C6EFCE")  # light green

    # summary block
    ws["A1"].value = "Summary"
    ws["A1"].font = header_font
    ws["A2"].value = "Total comparable rows"
    ws["B2"].value = int(summary.get("total_rows", 0))

    ws["A4"].value = "Strict accuracy"
    ws["A4"].font = header_font
    ws["A5"].value = "Year"
    ws["B5"].value = summary.get("strict_year", 0.0)
    ws["A6"].value = "Make"
    ws["B6"].value = summary.get("strict_make", 0.0)
    ws["A7"].value = "Model"
    ws["B7"].value = summary.get("strict_model", 0.0)
    ws["A8"].value = "Trim"
    ws["B8"].value = summary.get("strict_trim", 0.0)
    ws["A9"].value = "All 4 fields"
    ws["B9"].value = summary.get("strict_all", 0.0)

    ws["A11"].value = "Relaxed accuracy"
    ws["A11"].font = header_font
    ws["A12"].value = "Year"
    ws["B12"].value = summary.get("relaxed_year", 0.0)
    ws["A13"].value = "Make"
    ws["B13"].value = summary.get("relaxed_make", 0.0)
    ws["A14"].value = "Model"
    ws["B14"].value = summary.get("relaxed_model", 0.0)
    ws["A15"].value = "Trim"
    ws["B15"].value = summary.get("relaxed_trim", 0.0)
    ws["A16"].value = "All 4 fields"
    ws["B16"].value = summary.get("relaxed_all", 0.0)

    for cell in ["B5", "B6", "B7", "B8", "B9", "B12", "B13", "B14", "B15", "B16"]:
        ws[cell].number_format = "0.0%"

    # table header
    start_row = 18
    headers = [
        "Sheet1 Row",
        "Scraped Name",
        "QA Status",
        "Actual Year",
        "Expected Year",
        "Year (Strict)",
        "Actual Make",
        "Expected Make",
        "Make (Strict)",
        "Actual Model",
        "Expected Model",
        "Model (Strict)",
        "Actual Trim",
        "Expected Trim",
        "Trim (Strict)",
        "All 4 (Strict)",
        "All 4 (Relaxed)",
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=i)
        cell.value = h
        cell.font = header_font
        cell.alignment = center

    # rows
    for idx, r in enumerate(rows, start=1):
        rr = start_row + idx
        data = [
            r.row_sheet1,
            r.scraped_name,
            r.qa_status,
            r.actual_year,
            r.expected_year,
            "PASS" if r.strict_year else "FAIL",
            r.actual_make,
            r.expected_make,
            "PASS" if r.strict_make else "FAIL",
            r.actual_model,
            r.expected_model,
            "PASS" if r.strict_model else "FAIL",
            r.actual_trim,
            r.expected_trim,
            "PASS" if r.strict_trim else "FAIL",
            "PASS" if r.strict_all else "FAIL",
            "PASS" if r.relaxed_all else "FAIL",
        ]
        for c, v in enumerate(data, start=1):
            cell = ws.cell(row=rr, column=c)
            cell.value = v
            if c in {1, 3, 6, 9, 12, 15, 16, 17}:
                cell.alignment = center

        # highlight strict failures per field
        def _paint(col_idx: int, ok: bool) -> None:
            ws.cell(row=rr, column=col_idx).fill = good_fill if ok else bad_fill

        _paint(6, r.strict_year)
        _paint(9, r.strict_make)
        _paint(12, r.strict_model)
        _paint(15, r.strict_trim)
        _paint(16, r.strict_all)
        # relaxed overall
        ws.cell(row=rr, column=17).fill = good_fill if r.relaxed_all else bad_fill

    # simple column widths
    widths = {
        1: 10,
        2: 45,
        3: 12,
        4: 11,
        5: 11,
        6: 11,
        7: 18,
        8: 18,
        9: 11,
        10: 22,
        11: 22,
        12: 12,
        13: 18,
        14: 18,
        15: 11,
        16: 12,
        17: 13,
    }
    for col, w in widths.items():
        ws.column_dimensions[chr(ord("A") + col - 1)].width = w

    ws.freeze_panes = ws["A19"]
    wb.save(output_path)


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python qa_accuracy_report.py <input.xlsx> [output.xlsx]")
        return 2

    input_path = sys.argv[1]
    output_path = (
        sys.argv[2]
        if len(sys.argv) >= 3
        else re.sub(r"\.xlsx$", ".accuracy.xlsx", input_path, flags=re.IGNORECASE)
    )
    rows, summary = compare_workbook(input_path)

    total = int(summary.get("total_rows", 0))
    print(f"Total comparable rows: {total}")
    print("\nStrict accuracy:")
    print(f"  Year:  {summary['strict_year']:.1%}")
    print(f"  Make:  {summary['strict_make']:.1%}")
    print(f"  Model: {summary['strict_model']:.1%}")
    print(f"  Trim:  {summary['strict_trim']:.1%}")
    print(f"  ALL:   {summary['strict_all']:.1%}")

    print("\nRelaxed accuracy:")
    print(f"  Year:  {summary['relaxed_year']:.1%}")
    print(f"  Make:  {summary['relaxed_make']:.1%}")
    print(f"  Model: {summary['relaxed_model']:.1%}")
    print(f"  Trim:  {summary['relaxed_trim']:.1%}")
    print(f"  ALL:   {summary['relaxed_all']:.1%}")

    write_report(input_path, output_path, rows, summary)
    print(f"\nWrote report: {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


